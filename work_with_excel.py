from openpyxl import load_workbook  # https://openpyxl.readthedocs.io/en/stable/

# Загружаем существующий файл
wb = load_workbook("data/Списочный_состав.xlsx")
# Получаем активный лист (тот, который был открыт при сохранении)
ws = wb.active
# Читаем данные из ячейки
print(ws["F6"].value)

# Чтение строк с 6 по 189
for row in ws.iter_rows(min_row=6, max_row=189, values_only=True):
    print(row)
    print(row[5], row[6], row[3])